from datetime import datetime
from multiprocessing import context
from django.shortcuts import render

from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views import View
from django.urls import resolve
from django.http import HttpResponse

from django.db.models import Sum, OuterRef, Subquery, F, Value, Count, Case, When, FloatField
from django.db.models.functions import Round, Cast
from document.models import DocSKAI, MacroData

from monev.models import LRPA_Monitoring, LRPA_File, MouPengalihanData, FileMouPengalihan
from document.models import PRK, PRKData
from monev.utils import get_all_prk_last_lrpa, get_latest_rekom_period, get_last_lrpa
from monev.views import this_month, is_production

from .models import UsulanPeriod, UsulanRekomposisi, UsulanRekomposisiData, EbudgetFile
from .forms import UsulanPeriodForm

from num2words import num2words
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook

class RecompositionPeriod(LoginRequiredMixin, View):
    def get(self, request):
        context = {}

        period =  UsulanPeriod.objects.all()
        context["period"] = period

        return render(request, 'recomposition/recomposition_aki_period.html', context)

class RecompositionPeriodCreate(LoginRequiredMixin, View):
    def get(self, request):
        if request.user.is_admin:
            return render(request, 'recomposition/recomposition_period_create.html', {})
        return render(request, '403_forbidden.html', {})
    
    def post(self, request):
        print(request.POST)
        form = UsulanPeriodForm(request.POST)

        if form.is_valid():
            period = form.save(commit=False)
            period.created_by = request.user
            period.save()

            if period.is_before_period():
                period.status = -1
                period.save()
            
            # CHECK IF THIS IS PERIOD FOR REKOM AKI
            list_bpo = ["REN", "K3L", "OPK 1", "OPK 2", "PPK"]
            if period.for_rekom_aki:
                for bpo in list_bpo:
                    draft = UsulanRekomposisi(division=bpo,period=period)
                    draft.save()

                    try:
                        lrpa = get_all_prk_last_lrpa(rekap_user_induk=bpo)
                        if not draft.is_data_created:
                            for data in lrpa:
                                usulan_data = UsulanRekomposisiData(file=draft, prk=data.prk)
                                usulan_data.save()
                            
                            draft.is_data_created = True
                            draft.save()
                    except Exception as e:
                        print(e)

            # CHECK IF THIS IS PERIOD FOR REKOM AKB
            if period.for_rekom_aki:
                for bpo in list_bpo:
                    draft = UsulanRekomposisi(division=bpo,period=period,revisi="AKB")
                    draft.save()

                    try:
                        lrpa = get_all_prk_last_lrpa(rekap_user_induk=bpo)
                        if not draft.is_data_created:
                            for data in lrpa:
                                usulan_data = UsulanRekomposisiData(file=draft, prk=data.prk)
                                usulan_data.save()
                            
                            draft.is_data_created = True
                            draft.save()
                    except Exception as e:
                        print(e)
        else:
            print(form.errors)

        return render(request, 'recomposition/recomposition_period_create.html', {})

class RecompositionOutput(LoginRequiredMixin, View):
    def get(self, request):
        context = {}
        if is_production():
            skai_3 = DocSKAI.objects.get(pk=6)
        else:
            skai_3 = DocSKAI.objects.get(pk=19)

        combine_list = []
        all_prk_data = get_all_prk_last_lrpa()
        sq_3 = MacroData.objects.filter(macro_file=skai_3.macro.macro_file_1, prk=OuterRef('prk'))
        
        combine_query = all_prk_data.annotate(status=sq_3.values('ang_status'),sumber_dana = sq_3.values('sumber_dana'),
        aki_n1 = Round(Subquery(sq_3.values('aki_n1_year')[:1])),aki_n2 = Round(Subquery(sq_3.values('aki_n2_year')[:1])),
        aki_n3 = Round(Subquery(sq_3.values('aki_n1_year')[:1])),aki_n4 = Round(Subquery(sq_3.values('aki_n1_year')[:1])),
        aki_after_n1 = Round(Subquery(sq_3.values('aki_after_n1_year')[:1])))
        for data in combine_query:
            try:
                usulan_data = UsulanRekomposisiData.objects.get(file__period=get_latest_rekom_period(), prk=data.prk, file__revisi="AKI")
                is_changed = usulan_data.is_changed
                if is_changed:
                    temp_usulan = data.get_total_realisasi()
                    for i in range(this_month(),13):
                        if usulan_data.get_rencana_bulan(i):
                            temp_usulan += usulan_data.get_rencana_bulan(i)
                        else:
                            temp_usulan += int(float(data.get_rencana_month(i) or 0))
                    
                    temp_usulan -= data.get_current_month_realisasi()
                    
                    selisih_usulan = temp_usulan - data.real_aki()
                else:
                    temp_usulan = None
                    selisih_usulan = 0
            except:
                usulan_data = None
                temp_usulan = 0
                selisih_usulan = 0
                is_changed = None
            
            aki_rekom_tahun = 0 if data.status == "Murni" else data.alternate_aki()
            aki_belum_terbit = data.alternate_aki() - aki_rekom_tahun

            combine_list.append((data, usulan_data, temp_usulan, selisih_usulan, is_changed, aki_rekom_tahun, aki_belum_terbit))

        context["all_data"] = combine_list
        context["month"] = this_month()
        return render(request, 'recomposition/recomposition_aki_output.html', context)
    
    def post(self, request):
        context = {}
        file = EbudgetFile.objects.last()
        wb = load_workbook(file.file)
        ws = wb['Rekomposisi']
        list_rows = [idx for idx,cell in enumerate(ws["B"]) if cell.value and idx >= 9]

        #TRY ADD ONE LAST ROW #INTENTIONALLY MISSING LAST ROW
        # last_idx = list_rows[-1]
        # last_idx = last_idx + 1
        # list_rows.append(last_idx)
        start_col = 1 # B Column
        end_col = 37 # AL Column 
        for rows in list_rows:
            row_temp = [cell for cell in ws[rows][start_col:end_col+1]]
            row = [cell.value for cell in ws[rows][start_col:end_col+1]]
            cell_temp = row_temp[36]
            cell_temp.value = 5
            print(cell_temp.value)
        
        wb.save("doc.xlsx")
        print("DONE")

        return render(request, 'recomposition/snippets/button_download_xls.html', context)

class eBudgetOutput(LoginRequiredMixin, View):
    def get(self, request):
        context = {}
        file = EbudgetFile.objects.last()
        wb = load_workbook(file.file)
        ws = wb['Rekomposisi']
        list_rows = [idx for idx,cell in enumerate(ws["B"]) if cell.value and idx >= 9]

        #TRY ADD ONE LAST ROW #INTENTIONALLY MISSING LAST ROW
        # last_idx = list_rows[-1]
        # last_idx = last_idx + 1
        # list_rows.append(last_idx)
        start_col = 1 # B Column
        end_col = 37 # AL Column 
        for rows in list_rows:
            row_temp = [cell for cell in ws[rows][start_col:end_col+1]]
            row = [cell.value for cell in ws[rows][start_col:end_col+1]]
            cell_temp = row_temp[36]
            cell_temp.value = 5
            print(cell_temp.value)
        
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename=myexport.xlsx'
        return response

class RecompositionAKI(LoginRequiredMixin, View):

    def get(self, request):
        context = {}
        context["month"] = this_month()
        combine_list = []
        rekap_list = []

        # GET LATEST PERIOD
        period = get_latest_rekom_period()
        context["period"] = period
        
        # CHECK IF IN PERIOD
        if not period.is_in_period():
            return render(request, '404_not_found_2.html', context)
        
        if not period.for_rekom_aki:
            return render(request, '404_not_found_2.html', context)

        last_lrpa = get_last_lrpa()
        last_mou = FileMouPengalihan.objects.order_by('file_export_date').first()

        #GET USER DIVISION, DETERMINE USER VIEW
        division = request.user.division

        if division in ["ANG", "KEU", "Super Admin", "Admin - Intern"]:

            print(division)
            
            # GET ALL VIEW
            rekap_user = ["REN", "PPK", "OPK 1", "OPK 2", "K3L"]
            
            total_usulan = 0
            missing_notes = []
            
            for_div = ""

            for user in rekap_user:
                lrpa = get_all_prk_last_lrpa(rekap_user_induk=user)
                total_aki = lrpa.aggregate(Sum('aki_this_year'))['aki_this_year__sum']

                #GET LATEST DRAFT
                draft = UsulanRekomposisi.objects.get(division=user, period=period, revisi="AKI")
                query_data = UsulanRekomposisiData.objects.filter(file=draft, prk__rekap_user_induk=user)
                count_chg  = query_data.aggregate(chg=Count(Case(When(is_changed=True, then=1))))

                if draft.is_publish == False:
                    continue
                
                for_div += user
                for_div += " "

                #IF USULAN DATA NOT CREATED YET
                if not draft.is_data_created:
                    for data in lrpa:
                        usulan_data = UsulanRekomposisiData(file=draft, prk=data.prk)
                        usulan_data.save()
                    
                    draft.is_data_created = True
                    draft.save()
                
                #GET DATA INFO
                for data in lrpa:
                    usulan_data = UsulanRekomposisiData.objects.get(file=draft, prk=data.prk)
                    temp_usulan = data.get_total_realisasi()
                    for i in range(this_month(),13):
                        if usulan_data.get_rencana_bulan(i):
                            temp_usulan += usulan_data.get_rencana_bulan(i)
                        else:
                            temp_usulan += int(float(data.get_rencana_month(i) or 0))
                    
                    temp_usulan -= data.get_current_month_realisasi()
                    
                    selisih_usulan = temp_usulan - data.real_aki()

                    combine_list.append((data,usulan_data,temp_usulan,selisih_usulan))
                    total_usulan = total_usulan + temp_usulan
                    
                    if usulan_data.is_missing_notes():
                        msg_notes = "PRK : " + str(usulan_data.prk.no_prk) + " terdapat perubahan AKB tetapi tidak mencantumkan notes"
                        missing_notes.append((msg_notes))

                rekap_list.append((user,count_chg["chg"],draft.last_edit_date, draft.is_publish, draft.proposed_by, draft.upload_date, total_usulan, total_aki))

            
            
            context["for_div"] = for_div
            document = [last_lrpa, last_mou]
            context["document"] = document #OPTIMIZE LATER?

            context["lrpa"] = combine_list
            context["rekap_list"] = rekap_list


            return render(request, 'recomposition/recomposition_aki_admin.html', context)
        else:
           
            lrpa = PRKData.objects.select_related('prk').filter(file_lrpa=last_lrpa, prk__rekap_user_induk=division)
            context["for_div"] = division
        
            #GET LATEST DRAFT
            draft = UsulanRekomposisi.objects.get(division=context["for_div"], period=period, revisi="AKI")
            context["draft_pk"] = draft.pk

            # CHECK IF ALREADY SUBMITTED # CHANGE TEMPLATE LATER
            if draft.is_publish == True:
                return render(request, '404_not_found_2.html', context)

            # IF USULAN DATA NOT CREATED YET    
            if not draft.is_data_created:
                for data in lrpa:
                    usulan_data = UsulanRekomposisiData(file=draft, prk=data.prk)
                    usulan_data.save()
                
                draft.is_data_created = True
                draft.save()
        
            total_usulan = 0
            missing_notes = []

            for data in lrpa:
                # SELECT THROUGH THE MONTH
                usulan_data = UsulanRekomposisiData.objects.get(file=draft, prk=data.prk)
                temp_usulan = data.get_total_realisasi()
                for i in range(this_month(),13):
                    if usulan_data.get_rencana_bulan(i):
                        temp_usulan += usulan_data.get_rencana_bulan(i)
                    else:
                        temp_usulan += int(float(data.get_rencana_month(i) or 0))
                
                selisih_usulan = temp_usulan - data.real_aki()

                combine_list.append((data,usulan_data,temp_usulan,selisih_usulan))
                total_usulan = total_usulan + temp_usulan
                
                if usulan_data.is_missing_notes():
                    msg_notes = "PRK : " + str(usulan_data.prk.no_prk) + " terdapat perubahan AKB tetapi tidak mencantumkan notes"
                    missing_notes.append((msg_notes))
            
            if 'check-data' in request.GET:
                context["show_error"] = True

                total_aki = lrpa.aggregate(Sum('aki_this_year'))['aki_this_year__sum']
                error_count = 0
                #CHECK DEFISIT USULAN
                if total_aki != total_usulan:
                    msg = "Total AKI tidak sama dengan selisih usulan terdapat "
                    if total_aki < total_usulan:
                        msg = msg + "surplus sebesar Rp."
                        msg = msg + str(int(float(total_usulan-total_aki)))
                    
                    if total_aki > total_usulan:
                        msg = msg + "defisit sebesar Rp."
                        msg = msg + str(int(float(total_aki-total_usulan)))

                    context["aki_error"] = msg
                
                #CHECK IF THERE IS MISSING NOTES
                context["missing_notes"] = missing_notes
                error_count = error_count + len(missing_notes)
                context["error_count"] = error_count
                if error_count > 0:
                    context["error_count_msg"] = "Terdapat total " + str(error_count) + " error"
                
                if error_count == 0:
                    context["error_count_msg"] = "Sudah tidak terdapat error, data rekomposisi sudah bisa dikirim"

        document = [last_lrpa, last_mou]
        context["document"] = document #OPTIMIZE LATER?

        context["lrpa"] = combine_list

        return render(request, 'recomposition/recomposition_aki.html', context)
    
    def post(self, request):
        context = {}
        draft = UsulanRekomposisi.objects.get(pk=request.POST["draft-pk"])
        draft.is_publish = True
        draft.proposed_by = request.user
        draft.upload_date = datetime.now()
        draft.save()

        if request.user.division not in ["ANG", "KEU", "Super Admin"] and draft.is_publish == True:
            return render(request, '404_not_found_2.html', context)

        return render(request, 'recomposition/recomposition_aki.html', context)

class RecompositionAKB(LoginRequiredMixin, View):

    def get(self, request):
        context = {}
        context["month"] = this_month()
        combine_list = []
        rekap_list = []
        # GET LATEST PERIOD
        period = get_latest_rekom_period()
        context["period"] = period
        
        # CHECK IF IN PERIOD
        if not period.is_in_period():
            return render(request, '404_not_found_2.html', context)

        last_lrpa = get_last_lrpa()
        last_mou = FileMouPengalihan.objects.order_by('file_export_date').first()

        #GET USER DIVISION, DETERMINE USER VIEW
        division = request.user.division

        if division in ["ANG", "KEU", "Super Admin", "Admin - Intern"]:
            
            # ADMIN VIEW
            # GET ALL VIEW
            rekap_user = ["REN", "PPK", "OPK 1", "OPK 2", "K3L"]
            
            total_usulan = 0
            missing_notes = []
            
            for_div = ""

            for user in rekap_user:
                lrpa = PRKData.objects.select_related('prk').filter(file_lrpa=last_lrpa, prk__rekap_user_induk=user)
                total_aki = lrpa.aggregate(Sum('aki_this_year'))['aki_this_year__sum']
                #GET LATEST DRAFT
                draft = UsulanRekomposisi.objects.get(division=user, period=period, revisi="AKB")
                query_data = UsulanRekomposisiData.objects.filter(file=draft, prk__rekap_user_induk=user)
                count_chg  = query_data.aggregate(chg=Count(Case(When(is_changed=True, then=1))))

                if draft.is_publish == False:
                    continue
                
                for_div += user
                for_div += " "

                #IF USULAN DATA NOT CREATED YET
                if not draft.is_data_created:
                    for data in lrpa:
                        usulan_data = UsulanRekomposisiData(file=draft, prk=data.prk)
                        usulan_data.save()
                    
                    draft.is_data_created = True
                    draft.save()
                
                #GET DATA INFO
                for data in lrpa:
                    usulan_data = UsulanRekomposisiData.objects.get(file=draft, prk=data.prk)
                    temp_usulan = data.get_total_realisasi()
                    for i in range(this_month(),13):
                        if usulan_data.get_rencana_bulan(i):
                            temp_usulan += usulan_data.get_rencana_bulan(i)
                        else:
                            temp_usulan += int(float(data.get_rencana_month(i) or 0))
                        
                    temp_usulan -= data.get_current_month_realisasi()

                    selisih_usulan = temp_usulan - data.real_aki()

                    combine_list.append((data,usulan_data,temp_usulan,selisih_usulan))
                    total_usulan = total_usulan + temp_usulan
                    
                    if usulan_data.is_missing_notes():
                        msg_notes = "PRK : " + str(usulan_data.prk.no_prk) + " terdapat perubahan AKB tetapi tidak mencantumkan notes"
                        missing_notes.append((msg_notes))

            context["for_div"] = for_div
            document = [last_lrpa, last_mou]
            context["document"] = document #OPTIMIZE LATER?

            context["lrpa"] = combine_list

            return render(request, 'recomposition/recomposition_akb_admin.html', context)
        else:
           
            lrpa = PRKData.objects.select_related('prk').filter(file_lrpa=last_lrpa, prk__rekap_user_induk=division)
            context["for_div"] = division
        
            #GET LATEST DRAFT
            draft = UsulanRekomposisi.objects.get(division=context["for_div"], period=period, revisi="AKB")
            context["draft_pk"] = draft.pk

            # CHECK IF ALREADY SUBMITTED # CHANGE TEMPLATE LATER
            if draft.is_publish == True:
                return render(request, '404_not_found_2.html', context)

            # IF USULAN DATA NOT CREATED YET
            
            if not draft.is_data_created:
                for data in lrpa:
                    usulan_data = UsulanRekomposisiData(file=draft, prk=data.prk)
                    usulan_data.save()
                
                draft.is_data_created = True
                draft.save()
        
            total_usulan = 0
            missing_notes = []
            for data in lrpa:
                # SELECT THROUGH THE MONTH
                usulan_data = UsulanRekomposisiData.objects.get(file=draft, prk=data.prk)
                temp_usulan = data.get_total_realisasi()
                for i in range(this_month(),13):
                    if usulan_data.get_rencana_bulan(i):
                        temp_usulan += usulan_data.get_rencana_bulan(i)
                    else:
                        temp_usulan += int(float(data.get_rencana_month(i) or 0))
                
                selisih_usulan = temp_usulan - data.real_aki()
                print(data.prk.no_prk, selisih_usulan)

                combine_list.append((data,usulan_data,temp_usulan,selisih_usulan))
                total_usulan = total_usulan + temp_usulan
                
                if usulan_data.is_missing_notes():
                    msg_notes = "PRK : " + str(usulan_data.prk.no_prk) + " terdapat perubahan AKB tetapi tidak mencantumkan notes"
                    missing_notes.append((msg_notes))
            
            if 'check-data' in request.GET:
                context["show_error"] = True

                total_aki = lrpa.aggregate(Sum('aki_this_year'))['aki_this_year__sum']
                error_count = 0
                #CHECK DEFISIT USULAN
                if total_aki != total_usulan:
                    msg = "Total AKB tidak sama. Terdapat "
                    error_count += 1
                    if total_aki < total_usulan:
                        msg = msg + "surplus sebesar Rp."
                        msg = msg + str(int(float(total_usulan-total_aki)))
                    
                    if total_aki > total_usulan:
                        msg = msg + "defisit sebesar Rp."
                        msg = msg + str(int(float(total_aki-total_usulan)))

                    context["aki_error"] = msg
                
                #CHECK IF THERE IS MISSING NOTES
                context["missing_notes"] = missing_notes
                error_count += len(missing_notes)
                context["error_count"] = error_count
                if error_count > 0:
                    context["error_count_msg"] = "Terdapat total " + str(error_count) + " error"
                
                if error_count == 0:
                    context["error_count_msg"] = "Sudah tidak terdapat error, data rekomposisi sudah bisa dikirim"

        document = [last_lrpa, last_mou]
        context["document"] = document #OPTIMIZE LATER?

        context["lrpa"] = combine_list

        return render(request, 'recomposition/recomposition_akb.html', context)
    
    def post(self, request):
        context = {}
        draft = UsulanRekomposisi.objects.get(pk=request.POST["draft-pk"])
        draft.is_publish = True
        draft.proposed_by = request.user
        draft.upload_date = datetime.now()
        draft.save()

        if request.user.division not in ["ANG", "KEU", "Super Admin"] and draft.is_publish == True:
            return render(request, '404_not_found_2.html', context)

        return render(request, 'recomposition/recomposition_akb.html', context)

class RecompositionAKBRekap(LoginRequiredMixin, View):

    def get(self, request):
        context = {}
        rekap_list = []
        last_lrpa = get_last_lrpa()
         #GET REKAP TIAP BULAN
        month_name = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
        field_name = ["jan_rencana", "feb_rencana", "mar_rencana", "apr_rencana", "mei_rencana", "jun_rencana", "jul_rencana", "aug_rencana", "sep_rencana", "okt_rencana", "nov_rencana", "des_rencana"]
        lrpa_1 = PRKData.objects.select_related('prk').filter(file_lrpa=last_lrpa, mekanisme_pembayaran="Unit")
        lrpa_2 = PRKData.objects.select_related('prk').filter(file_lrpa=last_lrpa, mekanisme_pembayaran="Pengalihan")
        lrpa_3 = PRKData.objects.select_related('prk').filter(file_lrpa=last_lrpa, mekanisme_pembayaran="Pusat")
        
        unit_sebelum_total = pengalihan_sebelum_total = pusat_sebelum_total = 0

        for idx,m in enumerate(month_name):
            # unit_sebelum = lrpa_1.annotate(as_float=Cast(field_name[idx], FloatField())).aggregate(Sum('as_float'))['as_float__sum']
            # pengalihan_sebelum = lrpa_2.annotate(as_float=Cast(field_name[idx], FloatField())).aggregate(Sum('as_float'))['as_float__sum']
            # pusat_sebelum = lrpa_3.annotate(as_float=Cast(field_name[idx], FloatField())).aggregate(Sum('as_float'))['as_float__sum']

            unit_sebelum = sum([int(float(d.get_rencana_month(idx + 1) or 0)) for d in lrpa_1])
            pengalihan_sebelum = sum([int(float(d.get_rencana_month(idx + 1) or 0)) for d in lrpa_2])
            pusat_sebelum = sum([int(float(d.get_rencana_month(idx + 1) or 0)) for d in lrpa_3])

            unit_sebelum_total += unit_sebelum
            pengalihan_sebelum_total += pengalihan_sebelum
            pusat_sebelum_total += pusat_sebelum

            total_sebelum = unit_sebelum + pengalihan_sebelum + pusat_sebelum

            unit_sesudah = sum([int(float(d.get_rencana_month(idx + 1) or 0)) if not UsulanRekomposisiData.objects.get(file__revisi="AKB",prk=d.prk).get_rencana_bulan(idx + 1) else UsulanRekomposisiData.objects.get(file__revisi="AKB",prk=d.prk).get_rencana_bulan(idx + 1) for d in lrpa_1 ])
            pengalihan_sesudah = sum([int(float(d.get_rencana_month(idx + 1) or 0)) for d in lrpa_2])
            pusat_sesudah = sum([int(float(d.get_rencana_month(idx + 1) or 0)) for d in lrpa_3])

            rekap_list.append((m,unit_sebelum, pengalihan_sebelum, pusat_sebelum, total_sebelum, unit_sesudah, pengalihan_sesudah, pusat_sesudah))

        grand_total_sebelum = unit_sebelum_total + pengalihan_sebelum_total + pusat_sebelum_total
        grand_total = [unit_sebelum_total, pengalihan_sebelum_total, pusat_sebelum_total, grand_total_sebelum]
        context["rekap_list"] = rekap_list
        context["grand_total"] = grand_total

        return render(request, 'recomposition/recomposition_akb_rekap.html', context)

class UsulanRekomposisiEdit(LoginRequiredMixin, View):
    
    def get(self, request, pk, month, revisi, *args, **kwargs):
        context = {}
        months = {1:"Januari", 2:"Februari", 3:"Maret", 4:"April", 5:"Mei", 6:"Juni", 7:"Juli", 8:"Agustus", 9:"September", 10:"Oktober", 11:"November", 12:"Desember", 0:"pass"}

        prk = PRK.objects.get(pk=pk)
        context["data"] = prk

        # GET LATEST PERIOD
        period = UsulanPeriod.objects.order_by('-pk').first()
        context["period"] = period

        # CHECK IF IN PERIOD
        if not period.is_in_period():
            return render(request, '404_not_found_2.html', context)

        last_lrpa = LRPA_File.objects.order_by('-pk').first()
        prk_data = PRKData.objects.get(file_lrpa=last_lrpa, prk=prk)

        if month != 0:
            value = prk_data.get_rencana_month(month) #MANUAL
            try:
                context["words"] = num2words(int(value) , lang='id')
                context["value"] = str(int(value))
            except:
                context["words"] = num2words(int(float(value)) , lang='id')
                context["value"] = str(int(float(value)))
        else:
            value = None
        
        try:
            if revisi == 0:
                draft = UsulanRekomposisi.objects.get(division=request.user.division, period=period, revisi="AKI")
                context["revisi"] = 0
            else:
                draft = UsulanRekomposisi.objects.get(division=request.user.division, period=period, revisi="AKB")
                context["revisi"] = 1

            prk_draft = UsulanRekomposisiData.objects.get(file=draft, prk=prk)
            value_1 = prk_draft.get_rencana_bulan(month) #MANUAL
            notes = prk_draft.notes
        except:
            value_1 = 0
            notes = ""
        
        context["value_draft"] = str(int(float(value_1))) if value_1 else str(int(float(value or 0)))
        context["current_month"] = this_month()
        context["current_month_realisasi"] = str(prk_data.get_current_month_realisasi())
        context["this_month"] = month
        context["month"] = months[month]
        context["selisih"] = str(value) #MANUAL
        context["notes"] = notes
        
        try:
            context["is_month"] = prk_draft.is_rencana_bulan(month)
        except:
            context["is_month"] = 0

        return render(request, 'recomposition/snippets/modal_edit.html', context)

    def post(self, request, pk, type, *args, **kwargs):
        pass

class OnChangeValue(LoginRequiredMixin, View):

    def get(self, request, former_value):
        context = {}
        data = request.GET
        context["words"] = num2words(int(data["value"]) , lang='id')
        

        selisih = int(data["value"]) - int(former_value)

        if selisih > 0:
            context["selisih"] = str(selisih) + " (tambah)"
        elif selisih < 0:
            context["selisih"] = str(selisih) + " (kurang)"
        else:
            context["selisih"] = str(selisih)

        print(context["selisih"])

        return render(request, 'recomposition/snippets/value_to_words.html', context)

class InlineAKBEdit(LoginRequiredMixin, View):

    def get(self, request):
        pass

    def post(self, request):
        context = {}

        data = request.POST
        prk = PRK.objects.get(no_prk = data["no_prk"])
        
        context["data"] = prk
        context["this_month"] = data["this_month"]
        context["revisi"] = data["revisi"]

        print("DATA = " + data["revisi"])
        
        # GET LATEST PERIOD
        period = UsulanPeriod.objects.order_by('-pk').first()
        context["period"] = period

        # CHECK IF IN PERIOD
        if not period.is_in_period():
            return render(request, '404_not_found_2.html', context)

        # UsulanRekomposisi should be created this time    
        division = request.user.division

        if data["revisi"] == "0":
            draft = UsulanRekomposisi.objects.get(division=division, period = period, revisi="AKI")
        else:
            draft = UsulanRekomposisi.objects.get(division=division, period = period, revisi="AKB")
        data_rekom = UsulanRekomposisiData.objects.get(file=draft, prk=prk)

        if not "delete" in data:

            if "notes" in data:
                data_rekom.notes = data["notes"]
                data_rekom.edited_by = request.user
                data_rekom.save()

                if not data_rekom.is_changed:
                    data_rekom.is_changed = True
                    data_rekom.save()
                
                draft.last_edit_date = datetime.now()
                draft.save()
                
                context["notes"] = data["notes"]
            
            if "value" in data:
                data_rekom.insertToMonth(data["this_month"], int(data["value"]))
                data_rekom.edited_by = request.user
                data_rekom.save()

                if not data_rekom.is_changed:
                    data_rekom.is_changed = True
                    data_rekom.save()
                
                context["edit_akb"] = data["value"]

                draft.last_edit_date = datetime.now()
                draft.save()
        
        else:
            data_rekom.insertToMonth(data["this_month"], None)
            data_rekom.edited_by = request.user
            data_rekom.is_changed = False
            data_rekom.save()

        return render(request, 'recomposition/snippets/inline_edit_cell_akb.html', context)

class InlineAKBDelete(LoginRequiredMixin, View):

    def get(self, request):
        pass

    def post(self, request):
        context = {}

        data = request.POST
        prk = PRK.objects.get(no_prk = data["no_prk"])
        
        context["data"] = prk
        context["this_month"] = data["this_month"]
        
        # GET LATEST PERIOD
        period = UsulanPeriod.objects.order_by('-pk').first()
        context["period"] = period

        # CHECK IF IN PERIOD
        if not period.is_in_period():
            return render(request, '404_not_found_2.html', context)

        # UsulanRekomposisi should be created this time        
        division = request.user.division
        draft = UsulanRekomposisi.objects.get(division=division, period = period)
        data_rekom = UsulanRekomposisiData.objects.get(file=draft, prk=prk)
        
        data_rekom.insertToMonth(data["this_month"], None)
        data_rekom.edited_by = request.user
        data_rekom.is_changed = False
        data_rekom.save()

        context["edit_akb"] = data["hidden_value_former"]

        return render(request, 'recomposition/snippets/inline_edit_cell_akb.html', context)