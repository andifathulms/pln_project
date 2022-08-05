from channels.consumer import AsyncConsumer
from channels.db import database_sync_to_async
from channels.generic.websocket import AsyncJsonWebsocketConsumer

import concurrent.futures

from openpyxl import load_workbook

from document.models import DocSKAI

from random import randint
from time import sleep


class MacrosConsumer(AsyncConsumer):

    async def websocket_connect(self, event):
        
        # When Socket Connected
        print("MacrosConsumer Connected with : " + str(self.scope["user"]))
        await self.send({"type": "websocket.accept",})

        await self.send({"type": "websocket.send", "text":0})
        
    
    async def websocket_receive(self, event):

        #Called when we get message from front-end
        print("MacrosConsumer: receive: ", event)

        await perform_macros_to_database(18)

        await self.send({"type":"websocket.send","text":str(randint(0,100))})
        
    async def websocket_disconnect(self, event):

        # When Socket Disconnected
        print("MacrosConsumer Disconnected", event)

@database_sync_to_async
def perform_macros_to_database(pk):
    doc = DocSKAI.objects.get(pk=pk)

    print("Load Workbook")
    wb = load_workbook(doc.macro_doc, keep_vba=True, data_only=True, read_only=True)
    print("Done")

    ws = wb['LKAI IDR']

    start_col = 2
    end_col = 24

    #list_rows = [idx for idx,cell in enumerate(ws["c"]) if cell.value and idx >= 9]
    #list_rows.append(list_rows[-1]+1)

    row_list = []
    row_ = 1
    for row in ws.rows:
        #print(row_)
        col = 0
        for cell in row:
            col += 1
            if col == 3:
                if cell.value:
                    row_list.append(row_)
        col = 0
        row_ += 1
        if row_ <= 9:
            continue
        if row_ >= 6500:
            break
    
    # print(row_list)
    for rows in row_list:
        if ws[rows][2].value != None:
            print("\n")
            print("Row = " + str(rows))
            #row = [cell.value for cell in ws[rows][start_col:end_col+1]]
            row = [ws[rows][col].value for col in range(start_col, end_col + 1)]
            print(row)
            print("\n")
        else:
            print("continue : " + str(rows))
            continue
    # def print_row(row):

    #     try:
    #         nonlocal ws
    #         print(ws)

    #         start_col = 2
    #         end_col = 24
    #         if ws["C"][row].value != None:
    #             print("Row = " + str(row))
    #             row = [cell.value for cell in ws[row][start_col:end_col+1]]
    #             print(row)
    #         else:
    #             print("continue : " + str(row))
    #     except Exception as e:
    #         print(e)

    # list_rows = [idx for idx,cell in enumerate(ws["c"]) if cell.value and idx >= 9]
    # print(list_rows)

    # with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
    #     print("Executor")
    #     executor.map(print_row,list_rows)
    #     print("Done Executor")
        
    print("Done!!!")

