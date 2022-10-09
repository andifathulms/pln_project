import json
from urllib import request
from django.shortcuts import redirect, render
from django.urls import reverse, reverse_lazy
from django.db.models import OuterRef, Subquery
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.conf import settings
from django.views import View
from django.views.generic.edit import DeleteView
from django.http import HttpResponseRedirect

from .models import Document, DocSKAI, MacroFile, Macro, MacroData, PRK
from monev.models import Assigned_PRK, PRK_Lookup
from .forms import DocumentForm

from itertools import chain

from tablib import Dataset
from django.utils import timezone

from openpyxl import load_workbook

from django.forms import model_to_dict
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models import Model

from monev.views import is_production
from monev.models import LRPA_Monitoring, LRPA_File, FileMouPengalihan, MouPengalihanData

class SKAIListView(LoginRequiredMixin, View):
    def get(self, request, pk, *args, **kwargs):
        context = {}
        today = timezone.now()
        doc_skai_1 = DocSKAI.objects.select_related('document').filter(document__published_date__year=today.year, type="Penetapan")
        doc_skai_2 = DocSKAI.objects.select_related('document').filter(document__published_date__year=today.year, type="Usulan")

        context["doc_skai_1"] = sorted(chain(doc_skai_1), key=lambda x: x.document.published_date, reverse=False)
        context["doc_skai_2"] = sorted(chain(doc_skai_2), key=lambda x: x.document.published_date, reverse=False)
        context["doc_skai_3"] = sorted(chain(doc_skai_1, doc_skai_2), key=lambda x: x.document.published_date, reverse=False)

        if pk == 1:
            context["doc_skai"] = sorted(chain(doc_skai_1), key=lambda x: x.document.published_date, reverse=False)
            context["skai_verb"] = "Penetapan"
        elif pk == 2:
            context["doc_skai"] = sorted(chain(doc_skai_2), key=lambda x: x.document.published_date, reverse=False)
            context["skai_verb"] = "Usulan"
        elif pk == 3:
            context["doc_skai"] = sorted(chain(doc_skai_1, doc_skai_2), key=lambda x: x.document.published_date, reverse=False)
            context["skai_verb"] = ""
        else:
            return HttpResponseRedirect(reverse('not_found'))

        year = DocSKAI.objects.values("year").distinct()
        context['year'] = year

        context['the_year'] = today.year
        return render(request, 'document/list_skai.html', context)
    
    def post(self, request, *args, **kwargs):
        context = {}
        doc_skai_1 = DocSKAI.objects.select_related('document').filter(document__published_date__year=request.POST["year"], type="Penetapan")
        doc_skai_2 = DocSKAI.objects.select_related('document').filter(document__published_date__year=request.POST["year"], type="Usulan")
        
        context["doc_skai_1"] = sorted(chain(doc_skai_1), key=lambda x: x.document.published_date, reverse=False)
        context["doc_skai_2"] = sorted(chain(doc_skai_2), key=lambda x: x.document.published_date, reverse=False)
        context["doc_skai_3"] = sorted(chain(doc_skai_1, doc_skai_2), key=lambda x: x.document.published_date, reverse=False)

        year = DocSKAI.objects.values("year").distinct()
        context['year'] = year
        context['the_year'] = request.POST["year"]
        return render(request, 'document/list_skai.html', context)

class SKAIDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = DocSKAI
    template_name = 'document/skai_delete.html'
    success_url = reverse_lazy('document:doc-list-skai')

    def test_func(self):
       return True #BIG WARNING

class SKAIUpdateView(LoginRequiredMixin, View):
    def get(self, request, pk, *args, **kwargs):
        context = {}

        skai = DocSKAI.objects.get(pk=pk)
        context["skai"] = skai
        context["str_date"] = skai.document.published_date.strftime('%Y-%m-%d')
        print(skai.document.published_date.strftime('%Y-%m-%d'))

        return render(request, 'document/skai_edit.html', context)
    
    def post(self, request, *args, **kwargs):
        context = {}
        
        skai = DocSKAI.objects.get(pk=request.POST["pk"])
        doc = skai.document

        skai.year = request.POST["year"]
        skai.keyword = request.POST["keyword"]

        doc.document_number = request.POST["document_number"]
        doc.regarding = request.POST["regarding"]
        doc.published_date = request.POST["published_date"]
        doc.save()
        skai.document = doc
        skai.save()

        if "file" in request.FILES:
            doc.file = request.FILES["file"]
            doc.save()
            skai.document = doc
            skai.save()
        
        #Macro doc later

        return redirect('home')


class LKAIView(LoginRequiredMixin, View):
    def get(self, request, pk, *args, **kwargs):
        context = {}

        skai = DocSKAI.objects.get(pk=pk)
        macro = skai.macro
        macro_1 = macro.macro_file_1

        macro_data = MacroData.objects.filter(macro_file=macro_1)

        context["skai"] = skai
        context["macros"] = macro_data

        return render(request, 'document/lkai_view.html', context)

    def post(self, request, *args, **kwargs):
        pass

class SKAIComparison(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        context = {}
        #skai_1 = DocSKAI.objects.get(pk=8) #DEV
        skai_1 = DocSKAI.objects.get(pk=1) #PROD
        macro_1 = skai_1.macro.macro_file_1
        macro_data_1 = MacroData.objects.filter(macro_file=macro_1)

        #skai_2 = DocSKAI.objects.get(pk=19) #DEV
        skai_2 = DocSKAI.objects.get(pk=6) #PROD
        macro_2 = skai_2.macro.macro_file_1
        macro_data_2 = MacroData.objects.filter(macro_file=macro_2)

        print(len(macro_data_1))
        print(len(macro_data_2))

        list_1 = list(data.no_prk for data in macro_data_1)
        list_2 = list(data.no_prk for data in macro_data_2)

        print(len(list(set(list_2)-set(list_1))))
        print(list(set(list_2)-set(list_1)))
        print(list(set(list_1)-set(list_2)))

        combine_list = []
        #ALL SKIP
        for data in macro_data_1:
            try:
                temp = MacroData.objects.get(no_prk=data.no_prk, macro_file=macro_2)
                combine_list.append((data,temp))
            except Exception as e:
                print("Skip " + str(data.no_prk))
                print(e)
            
        
        print(combine_list[0])
        print(combine_list[1])


        return render(request, 'document/skai_comparison.html', context)


class UploadSKAI(LoginRequiredMixin, View):

    def get(self, request, *args, **kwargs):
        context = {}
        doc_skai = DocSKAI.objects.all()
        context["doc_skai"] = doc_skai
        return render(request, 'document/upload_SKAI.html', context)

    
    def post(self, request, *args, **kwargs):
        context = {}
        
        if 'submit-skai' in request.POST:
            doc_form = DocumentForm(request.POST, request.FILES)
            if doc_form.is_valid():
                doc = doc_form.save(commit=False)
                doc.uploader = request.user
                doc.save()

                skai = DocSKAI(document=doc, year=request.POST['year'],keyword=request.POST['keyword'],revision=False,macro_doc=request.FILES["file_xls"])
                skai.save()

                #skai.create_notif_on_upload(skai.document.uploader,skai.document.regarding)
            else:
                print(doc.errors)
        elif 'submit-skai-usulan' in request.POST:
            doc_form = DocumentForm(request.POST, request.FILES)
            if doc_form.is_valid():
                doc = doc_form.save(commit=False)
                doc.uploader = request.user
                doc.save()

                skai = DocSKAI(document=doc, year=request.POST['year'],keyword=request.POST['keyword'],type="Usulan",macro_doc=request.FILES["file_xls"])
                skai.save()
                #skai.create_notif_on_upload(skai.document.uploader,skai.document.regarding)
        
        doc_skai = DocSKAI.objects.all()
        context["doc_skai"] = list(chain(doc_skai))

        return render(request, 'document/upload_SKAI.html', context)

class XLSM_Playground(UserPassesTestMixin, View):

    def test_func(self):
        return self.request.user.is_admin

    def get(self, request, *args, **kwargs):

        #MISSING ONE LAST ROW!!!!!! ??
        context = {}

        doc = DocSKAI.objects.get(pk=10)
        context["macros"] = doc.macro_doc
        print(doc.macro_doc)
        print("Load Workbook")
        wb = load_workbook(doc.macro_doc, keep_vba=True, data_only=True)
        print("Done")
        print(wb.sheetnames)
        ws = wb['LKAI IDR']
        ws_2 = wb['LKAI IDR 2']

        start_col = 31
        end_col = 58

        start_col_2 = 1
        end_col_2 = 40

        list_rows = [idx for idx,cell in enumerate(ws["AF"]) if cell.value and idx >= 9]

        #TRY ADD ONE LAST ROW
        last_idx = list_rows[-1]
        last_idx = last_idx + 1
        list_rows.append(last_idx)

        print(list_rows)
        print(ws['C'][9].value)
        macro_file = MacroFile()
        macro_file.save()

        exc = []
        for rows in list_rows:
            if ws["AF"][rows].value != None:
                
                print("Row = " + str(rows))
                row = [cell.value for cell in ws[rows][start_col:end_col+1]]
                row_2 = [cell.value for cell in ws_2[rows][start_col_2:end_col_2+1]]
                print(row)
                print(row_2)

                try:
                    macro_data = MacroData(macro_file = macro_file,
                        no_prk = row[0],
                        no_program = row[1],
                        no_ruptl = row[2],
                        cluster = row[3],
                        fungsi = row[4],
                        sub_fungsi = row[5],
                        program_utama = row[6],
                        score = row[7],
                        jenis_program = row[8],
                        keg_no = row[9],
                        keg_uraian = row[10],
                        keg_target_fisik = row[11],
                        keg_satuan = row[12],
                        ang_nilai = row[13],
                        ang_status = row[14],
                        ang_jenis_kontrak = row[15],
                        ang_no_kontrak = row[16],
                        realisasi_pembayaran = row[17],
                        prediksi_pembayaran = row[18],
                        ai_this_year = row[19],
                        aki_this_year = row[20],
                        aki_n1_year = row[21],
                        aki_n2_year = row[22],
                        aki_n3_year = row[23],
                        aki_n4_year = row[24],
                        aki_after_n1_year = row[25],
                        sumber_dana = row[26],
                        rencana_terkontrak = row_2[14],
                        rencana_COD = row_2[15],
                        jan_progress_fisik   = row_2[16],
                        jan_rencana_disburse = row_2[17], 
                        feb_progress_fisik   = row_2[18], 
                        feb_rencana_disburse = row_2[19], 
                        mar_progress_fisik   = row_2[20], 
                        mar_rencana_disburse = row_2[21], 
                        apr_progress_fisik   = row_2[22], 
                        apr_rencana_disburse = row_2[23], 
                        mei_progress_fisik   = row_2[24], 
                        mei_rencana_disburse = row_2[25], 
                        jun_progress_fisik   = row_2[26], 
                        jun_rencana_disburse = row_2[27], 
                        jul_progress_fisik   = row_2[28], 
                        jul_rencana_disburse = row_2[29], 
                        aug_progress_fisik   = row_2[30], 
                        aug_rencana_disburse = row_2[31], 
                        sep_progress_fisik   = row_2[32], 
                        sep_rencana_disburse = row_2[33], 
                        okt_progress_fisik   = row_2[34], 
                        okt_rencana_disburse = row_2[35], 
                        nov_progress_fisik   = row_2[36], 
                        nov_rencana_disburse = row_2[37], 
                        des_progress_fisik   = row_2[38], 
                        des_rencana_disburse = row_2[39] 
                    )
                    macro_data.save()

                except Exception as e:
                    exc.append(e)
                    print(e)                
            
            else:
                print("continue : " + str(rows))
                continue
        
        macro = Macro(macro_file_1=macro_file)
        macro.save()

        print("Done!!!")
        print(len(list_rows))
        print(exc)

        return render(request, 'document/playground.html', context)

class ExtendedEncoder(DjangoJSONEncoder):

    def default(self, o):

        if isinstance(o, Model):
            return model_to_dict(o)

        return super().default(o)

class JSON_Dumps(UserPassesTestMixin, View):

    def test_func(self):
        return self.request.user.is_admin

    def get(self, request, *args, **kwargs):

        context = {}

        # # TO DUMP JSON #
        doc = DocSKAI.objects.get(pk=3)
        # macro = doc.macro
        # macro_1 = macro.macro_file_1

        # macro_data = MacroData.objects.filter(macro_file=macro_1)
        # print(len(macro_data))

        # BASE_DIR = settings.BASE_DIR
        # # TO DUMP JSON #

        json_file = doc.json
        data1 = json.load(json_file) # deserialises it
        json_file.close()

        macro_file = MacroFile()
        macro_file.save()

        context["data"] = data1

        error = []

        for d in data1:
            try:
                macro_data = MacroData(macro_file = macro_file,
                    no_prk = d["no_prk"],
                    no_program = d["no_program"],
                    no_ruptl = d["no_ruptl"],
                    cluster = d["cluster"],
                    fungsi = d["fungsi"],
                    sub_fungsi = d["sub_fungsi"],
                    program_utama = d["program_utama"],
                    score = d["score"],
                    jenis_program = d["jenis_program"],
                    keg_no = d["keg_no"],
                    keg_uraian = d["keg_uraian"],
                    keg_target_fisik = d["keg_target_fisik"],
                    keg_satuan = d["keg_satuan"],
                    ang_nilai = d["ang_nilai"],
                    ang_status = d["ang_status"],
                    ang_jenis_kontrak = d["ang_jenis_kontrak"],
                    ang_no_kontrak = d["ang_no_kontrak"],
                    realisasi_pembayaran = d["realisasi_pembayaran"],
                    prediksi_pembayaran = d["prediksi_pembayaran"],
                    ai_this_year = d["ai_this_year"],
                    aki_this_year = d["aki_this_year"],
                    aki_n1_year = d["aki_n1_year"],
                    aki_n2_year = d["aki_n2_year"],
                    aki_n3_year = d["aki_n3_year"],
                    aki_n4_year = d["aki_n4_year"],
                    aki_after_n1_year = d["aki_after_n1_year"],
                    sumber_dana = d["sumber_dana"],
                    rencana_terkontrak = d["rencana_terkontrak"],
                    rencana_COD = d["rencana_COD"],
                    jan_progress_fisik   = d["jan_progress_fisik"],
                    jan_rencana_disburse = d["jan_rencana_disburse"], 
                    feb_progress_fisik   = d["feb_progress_fisik"], 
                    feb_rencana_disburse = d["feb_rencana_disburse"], 
                    mar_progress_fisik   = d["mar_progress_fisik"], 
                    mar_rencana_disburse = d["mar_rencana_disburse"], 
                    apr_progress_fisik   = d["apr_progress_fisik"], 
                    apr_rencana_disburse = d["apr_rencana_disburse"], 
                    mei_progress_fisik   = d["mei_progress_fisik"], 
                    mei_rencana_disburse = d["mei_rencana_disburse"], 
                    jun_progress_fisik   = d["jun_progress_fisik"], 
                    jun_rencana_disburse = d["jun_rencana_disburse"], 
                    jul_progress_fisik   = d["jul_progress_fisik"], 
                    jul_rencana_disburse = d["jul_rencana_disburse"], 
                    aug_progress_fisik   = d["aug_progress_fisik"], 
                    aug_rencana_disburse = d["aug_rencana_disburse"], 
                    sep_progress_fisik   = d["sep_progress_fisik"], 
                    sep_rencana_disburse = d["sep_rencana_disburse"], 
                    okt_progress_fisik   = d["okt_progress_fisik"], 
                    okt_rencana_disburse = d["okt_rencana_disburse"], 
                    nov_progress_fisik   = d["nov_progress_fisik"], 
                    nov_rencana_disburse = d["nov_rencana_disburse"], 
                    des_progress_fisik   = d["des_progress_fisik"], 
                    des_rencana_disburse = d["des_rencana_disburse"] 
                )
                macro_data.save()

            except Exception as e:
                print(e)
                error.append(e)
        
        context["error"] = error
        
        macro = Macro(macro_file_1=macro_file)
        macro.save()

        # # TO DUMP JSON #
        # with open("skai_revisi_1_new.json", 'w', encoding='utf-8') as outfile:
        #     for data in macro_data:
        #         d = json.dumps(data, cls=ExtendedEncoder, indent=4, separators=(',', ': '))
        #         outfile.write(',')
        #         try:
        #             outfile.write(d)
        #             print("DONE")
        #         except Exception as e:
        #             print(e)
        #         #print(d)
        
        # # TO DUMP JSON #

        return render(request, 'document/json_dumps.html', context)

class Assign_PRK(UserPassesTestMixin, View):

    def test_func(self):
        return self.request.user.is_admin

    def get(self, request, *args, **kwargs):
        context = {}
        file = Assigned_PRK.objects.get(pk=1)
        wb = load_workbook(file.file)
        ws = wb['Sheet1']
        start_col = 0
        end_col = 5
        list_rows = [idx for idx,cell in enumerate(ws["A"]) if cell.value and idx >= 1]
        print(list_rows)
        exc = []
        for rows in list_rows:
            if ws["A"][rows].value != None:
                
                print("Row = " + str(rows))
                row = [cell.value for cell in ws[rows][start_col:end_col+1]]
                print(row)

                try:
                    lookup = PRK_Lookup(file = file,
                        no_prk = row[0],
                        kode_prk = row[1],
                        kode_bpo = row[2],
                        upp = row[3],
                        rekap_user_induk = row[4]
                    )
                    lookup.save()

                except Exception as e:
                    exc.append(e)
                    print(e)                
            
            else:
                print("continue : " + str(rows))
                continue

        return render(request, 'document/json_dumps.html', context)
        
    def post(self, request, *args, **kwargs):
        pass

class PRKObject(UserPassesTestMixin, View):

    def test_func(self):
        return self.request.user.is_admin
    
    def get(self, request):
        context = {}

        # ALL DOCUMENT NEEDED
        if is_production():
            skai_1 = DocSKAI.objects.get(pk=1)
            skai_2 = DocSKAI.objects.get(pk=3)
            skai_3 = DocSKAI.objects.get(pk=6)
        else:
            skai_1 = DocSKAI.objects.get(pk=8)
            skai_2 = DocSKAI.objects.get(pk=10)
            skai_3 = DocSKAI.objects.get(pk=19)
        
        last_lrpa = LRPA_File.objects.order_by('-file_export_date').first()
        last_mou = FileMouPengalihan.objects.order_by('file_export_date').first()
        
        # macro_data_1 = MacroData.objects.filter(macro_file=skai_1.macro.macro_file_1).annotate(ai=Subquery(LRPA_Monitoring.objects.filter(file=last_lrpa,prk=OuterRef('prk')).values("ai_this_year")[:1]))
        # macro_data_2 = MacroData.objects.filter(macro_file=skai_2.macro.macro_file_1).order_by('no_prk')
        # macro_data_3 = MacroData.objects.filter(macro_file=skai_3.macro.macro_file_1).order_by('no_prk')

        # macros = [macro_data_1, macro_data_2, macro_data_3]

        sq_1 = MacroData.objects.filter(macro_file=skai_1.macro.macro_file_1, prk=OuterRef('prk'))
        sq_2 = MacroData.objects.filter(macro_file=skai_2.macro.macro_file_1, prk=OuterRef('prk'))
        sq_3 = MacroData.objects.filter(macro_file=skai_3.macro.macro_file_1, prk=OuterRef('prk'))
        sq_mou = MouPengalihanData.objects.filter(file=last_mou, prk=OuterRef('prk'))
        
        lrpa = LRPA_Monitoring.objects.select_related('prk').filter(file=last_lrpa). \
               annotate(ai_1 = sq_1.values('ai_this_year'), aki_1 = sq_1.values('aki_this_year'), status_1 = sq_1.values('ang_status'),
               ai_2 = sq_2.values('ai_this_year'), aki_2 = sq_2.values('aki_this_year'), status_2 = sq_2.values('ang_status'),
               ai_3 = sq_3.values('ai_this_year'), aki_3 = sq_3.values('aki_this_year'), status_3 = sq_3.values('ang_status'),
               mou_jan = sq_mou.values('jan'),mou_feb = sq_mou.values('feb'),mou_mar = sq_mou.values('mar'),mou_apr = sq_mou.values('apr'),
               mou_mei = sq_mou.values('mei'),mou_jun = sq_mou.values('jun'),mou_jul = sq_mou.values('jul'),mou_aug = sq_mou.values('aug'),
               mou_sep = sq_mou.values('sep'),mou_okt = sq_mou.values('okt'),mou_nov = sq_mou.values('nov'),mou_des = sq_mou.values('des')
               )
        for data in lrpa.iterator():
            print(data.no_prk, data.mou_jan, data.mou_feb, data.mou_mar, data.mou_apr)
            print(data.no_prk, data.mou_mei, data.mou_jun, data.mou_jul, data.mou_aug)
            print(data.no_prk, data.mou_sep, data.mou_okt, data.mou_nov, data.mou_des)
            print("###")
        

        #CREATING OBJECT PRK
        
        # for data in macro_data_1:
        #     if data.no_prk:
        #         try:
        #             obj = PRK.objects.get(no_prk=data.no_prk)
        #             print(data.no_prk, "Already There")
        #         except PRK.DoesNotExist:
        #             obj = PRK(
        #                 no_prk=data.no_prk,
        #                 no_program=data.no_program,
        #                 no_ruptl=data.no_ruptl,
        #                 cluster=data.cluster,
        #                 fungsi=data.fungsi,
        #                 sub_fungsi=data.sub_fungsi,
        #                 program_utama=data.program_utama,
        #                 score=data.score,
        #                 jenis_program=data.jenis_program,
        #                 keg_no=data.keg_no,
        #                 keg_uraian=data.keg_uraian
        #             )
        #             obj.save()
        #             print(data.no_prk, "Created")
        # print("##########################")
        # for data in macro_data_2:
        #     if data.no_prk:
        #         try:
        #             obj = PRK.objects.get(no_prk=data.no_prk)
        #             print(data.no_prk, "Already There")
        #         except PRK.DoesNotExist:
        #             obj = PRK(
        #                 no_prk=data.no_prk,
        #                 no_program=data.no_program,
        #                 no_ruptl=data.no_ruptl,
        #                 cluster=data.cluster,
        #                 fungsi=data.fungsi,
        #                 sub_fungsi=data.sub_fungsi,
        #                 program_utama=data.program_utama,
        #                 score=data.score,
        #                 jenis_program=data.jenis_program,
        #                 keg_no=data.keg_no,
        #                 keg_uraian=data.keg_uraian
        #             )
        #             obj.save()
        #             print(data.no_prk, "Created")
        # print("##########################")
        # for data in macro_data_3:
        #     if data.no_prk:
        #         try:
        #             obj = PRK.objects.get(no_prk=data.no_prk)
        #             print(data.no_prk, "Already There")
        #         except PRK.DoesNotExist:
        #             obj = PRK(
        #                 no_prk=data.no_prk,
        #                 no_program=data.no_program,
        #                 no_ruptl=data.no_ruptl,
        #                 cluster=data.cluster,
        #                 fungsi=data.fungsi,
        #                 sub_fungsi=data.sub_fungsi,
        #                 program_utama=data.program_utama,
        #                 score=data.score,
        #                 jenis_program=data.jenis_program,
        #                 keg_no=data.keg_no,
        #                 keg_uraian=data.keg_uraian
        #             )
        #             obj.save()
        #             print(data.no_prk, "Created")

        #ASSIGN PRK KODE TO PRK
        # for p in PRK.objects.all():
            
        #     try:
        #         lookup = PRK_Lookup.objects.get(no_prk=p.no_prk)
        #         p.kode_prk = lookup.kode_prk
        #         p.kode_bpo = lookup.kode_bpo
        #         p.rekap_user_induk = lookup.rekap_user_induk
        #         p.upp = lookup.upp

        #         p.save()

        #         print(p.no_prk, "Success")
        #     except Exception as e:
        #         print(p.no_prk, "None", e)

        #ASSIGN PRK OBJECT TO EACH MACRO DATA
        
        # # MACRO DATA
        # print("###################")
        # print("START ASSIGN MACROS")
        # print("###################")
        # for macro in macros:
        #     for data in macro:
        #         try:
        #             prk = PRK.objects.get(no_prk=data.no_prk)
        #             data.prk = prk
        #             data.save()

        #             print(data.prk.no_prk, "Success", data)
        #         except Exception as e:
        #             print(data.no_prk, "Failed", e, data)
        
        # print("#################")
        # print("END ASSIGN MACROS")
        # print("#################")
        
        # #LRPA

        # print("###################")
        # print("START ASSIGN LRPA")
        # print("###################")
        # last_lrpa = LRPA_File.objects.order_by('-file_export_date').first()
        # lrpa_data = LRPA_Monitoring.objects.filter(file=last_lrpa)
        # for data in lrpa_data:
        #     try:
        #         prk = PRK.objects.get(no_prk=data.no_prk)
        #         data.prk = prk
        #         data.save()

        #         print(data.prk.no_prk, "Success", data)
        #     except Exception as e:
        #         print(data.no_prk, "Failed", e, data)
        
        # print("#################")
        # print("END ASSIGN LRPA")
        # print("#################")

        # #MOU

        # print("###################")
        # print("START ASSIGN MOU")
        # print("###################")
        # last_mou = FileMouPengalihan.objects.order_by('file_export_date').first()
        # mou_data = MouPengalihanData.objects.filter(file=last_mou)
        # for data in mou_data:
        #     try:
        #         prk = PRK.objects.get(no_prk=data.no_prk)
        #         data.prk = prk
        #         data.save()

        #         print(data.prk.no_prk, "Success", data)
        #     except Exception as e:
        #         print(data.no_prk, "Failed", e, data)
        
        # print("#################")
        # print("END ASSIGN MOU")
        # print("#################")
        return render(request, 'document/json_dumps.html', context)