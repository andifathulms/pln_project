from django.shortcuts import render
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views import View
from django.db.models import Sum, OuterRef, Subquery, F, Value
from django.db.models.functions import Round

from document.models import DocSKAI, MacroData, PRK, PRKData

from openpyxl import load_workbook

from .models import LRPA_Monitoring, LRPA_File, PRK_Lookup, Assigned_PRK, MouPengalihanData, FileMouPengalihan
from .forms import LRPAFileForm, MouFileForm
from .utils import safe_div, this_month, is_production, get_last_lrpa, get_all_prk_last_lrpa

import sys

class DashboardView(LoginRequiredMixin, View):
    def get(self, request):
        context = {}
        months = {1:"Januari", 2:"Februari", 3:"Maret", 4:"April", 5:"Mei", 6:"Juni", 7:"Juli", 8:"Agustus", 9:"September", 10:"Oktober", 11:"November", 12:"Desember"}
        context["month"] = months[this_month()]

        last_lrpa = LRPA_File.objects.order_by('-pk').first()

        total_ai = 0
        total_aki = 0
        total_realisasi = 0
        pembayaran = ["Unit", "Pusat", "Pengalihan"]
        data_1 = []
        data_1_total = []
        for idx,x in enumerate(pembayaran):
            sum_ai = int(PRKData.objects.filter(file_lrpa=last_lrpa,mekanisme_pembayaran=x).aggregate(Sum('ai_this_year'))['ai_this_year__sum'])
            sum_aki = int(PRKData.objects.filter(file_lrpa=last_lrpa,mekanisme_pembayaran=x).aggregate(Sum('aki_this_year'))['aki_this_year__sum'])
            sum_realisasi = int(sum([m.get_total_realisasi() for m in PRKData.objects.filter(file_lrpa=last_lrpa,mekanisme_pembayaran=x)]))
            pct = (sum_realisasi*100)/sum_aki
            sisa = sum_aki - sum_realisasi

            total_ai = total_ai + sum_ai
            total_aki = total_aki + sum_aki
            total_realisasi = total_realisasi + sum_realisasi
            
            data_1.append((pembayaran[idx],sum_ai,sum_aki,sum_realisasi,pct,sisa))
        data_1_total.append((total_ai, total_aki,total_realisasi,(safe_div(total_realisasi,total_aki)*100),total_aki-total_realisasi))
        context["data_1"] = data_1
        context["data_1_total"] = data_1_total
        
        bpo = ["REN", "PPK", "OPK 1", "OPK 2", "K3L"]
        data_2 = []
        data_2_total = []
        
        total_akb = 0
        total_realisasi_bulan = 0
        for idx,x in enumerate(bpo):
            a = [int(float(p.get_rencana_month(this_month()))) for p in PRKData.objects.filter(file_lrpa=last_lrpa, prk__rekap_user_induk=x)]
            b = [int(float(p.get_realisasi_month(this_month()))) for p in PRKData.objects.filter(file_lrpa=last_lrpa, prk__rekap_user_induk=x)]
            c = [p.prk.no_prk for p in PRKData.objects.filter(file_lrpa=last_lrpa, prk__rekap_user_induk=x)]
            print(x,len(a), len(b), len(c))
            sum_of_akb = int(sum(a))
            sum_of_realisasi = int(sum(b))
            sisa = sum_of_akb - sum_of_realisasi
            try:
                pct = (sisa*100)/sum_of_akb
            except:
                pct = 0
            data_2.append((x, sum_of_akb, sum_of_realisasi,sisa,pct))
            a_copy = [value for value in a if value != 0]
            print(x, len(a_copy), a_copy)
            total_akb = total_akb + sum_of_akb
            total_realisasi_bulan = total_realisasi_bulan + sum_of_realisasi
        
        data_2_total.append((total_akb, total_realisasi_bulan,"-",total_akb-total_realisasi_bulan,(safe_div((total_akb-total_realisasi_bulan),total_akb)*100)))
        context["data_2"] = data_2
        context["data_2_total"] = data_2_total
        
        return render(request, 'account/dashboard.html', context)

class MonevView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        context = {}
        context["month"] = this_month()
        last_lrpa = LRPA_File.objects.order_by('-pk').first()
        last_mou = FileMouPengalihan.objects.order_by('file_export_date').first()

        month = this_month() #CHANGE LATER TO LAST MOU MONTH

        total_ai_bpo = 0
        total_aki_bpo = 0
        total_realisasi_bpo = 0

        #START COMPUTE FOR MONEV BY BPO
        #COUNT FOR BPO BESIDE "UPP"
        BPO_1 = ["Perencanaan", "Perizinan, Pertanahan, dan Komunikasi", "Operasi Konstruksi 1", "Operasi Konstruksi 2", "K3L"]
        BPO_2 = ["REN", "PPK", "OPK 1", "OPK 2", "K3L"]

        BPO_list = []

        for idx,data in enumerate(BPO_2):
            query = PRKData.objects.filter(prk__kode_bpo=data, file_lrpa=last_lrpa)
            count = query.count()
            sum_ai = query.aggregate(Sum('ai_this_year'))['ai_this_year__sum']
            sum_aki = query.aggregate(Sum('aki_this_year'))['aki_this_year__sum']
            realisasi = 0
            for x in query:
                realisasi = realisasi + x.get_total_realisasi()
            
            total_ai_bpo = total_ai_bpo + sum_ai
            total_aki_bpo = total_aki_bpo + sum_aki
            total_realisasi_bpo = total_realisasi_bpo + realisasi

            sisa_aki = sum_aki - realisasi
            pct = (realisasi/sum_aki)*100
            
            BPO_list.append((BPO_1[idx],data,sum_ai,sum_aki,realisasi,sisa_aki,pct,count))
        
        #COUNT FOR BPO "UPP"
        BPO_UPP_1 = ["UPP KITRING SULSEL","UPP KITRING SULTENG","UPP KITRING SULTRA","UPP KITRING SULUT dan GORONTALO"]
        BPO_UPP_2 = ["UPP 1","UPP 2", "UPP 3", "UPP 4"]

        for idx,data in enumerate(BPO_UPP_2):
            query = PRKData.objects.filter(prk__upp=data, file_lrpa=last_lrpa)
            count = query.count()
            sum_ai = query.aggregate(Sum('ai_this_year'))['ai_this_year__sum']
            sum_aki = query.aggregate(Sum('aki_this_year'))['aki_this_year__sum']
            if not sum_ai: sum_ai = 0
            if not sum_aki: sum_aki = 0
            realisasi = 0
            for x in query:
                realisasi = realisasi + x.get_total_realisasi()
            
            total_ai_bpo = total_ai_bpo + sum_ai
            total_aki_bpo = total_aki_bpo + sum_aki
            total_realisasi_bpo = total_realisasi_bpo + realisasi

            sisa_aki = sum_aki - realisasi
            pct = (safe_div(realisasi,sum_aki))*100
            
            BPO_list.append((BPO_UPP_1[idx],data,sum_ai,sum_aki,realisasi,sisa_aki,pct,count))
        
        context["BPO_list"] = BPO_list
        context["count_bpo_total"] = PRKData.objects.filter(prk__kode_bpo__in=BPO_2, file_lrpa=last_lrpa).count() + PRKData.objects.filter(prk__upp__in=BPO_UPP_2, file_lrpa=last_lrpa).count()
        context["total_ai_bpo"] = total_ai_bpo
        context["total_aki_bpo"] = total_aki_bpo
        context["total_realisasi_bpo"] = total_realisasi_bpo
        context["total_sisa_aki_bpo"] = total_aki_bpo - total_realisasi_bpo
        context["total_pct_bpo"] = (safe_div(total_realisasi_bpo,total_aki_bpo))*100
        
        #############################################

        #START COMPUTE FOR MONEV BY PRK
        #COUNT FOR ALL PRK IN "Pekerjaan. Prasarana"
        A_PRK_1 = ["Survey dan Soil Investigasi", "Perijinan", "Studi Lingkungan/AMDAL/UKL-UPL/LARAP", "Jasa Konsultasi Pembebasan tanah & ROW", "Inventarisasi, Pembebasan tanah & ROW"]
        A_PRK_2 = ["SV", "IZ", "LH", "JP", "TN"]

        A_list = []

        total_ai_a = 0
        total_aki_a = 0
        total_realisasi_a = 0

        count_a_total = 0

        for idx,data in enumerate(A_PRK_2):
            query = PRKData.objects.filter(prk__kode_prk=data, file_lrpa=last_lrpa)
            count = query.count()
            sum_ai = query.aggregate(Sum('ai_this_year'))['ai_this_year__sum']
            sum_aki = query.aggregate(Sum('aki_this_year'))['aki_this_year__sum']
            if not sum_ai: sum_ai = 0
            if not sum_aki: sum_aki = 0
            realisasi = 0
            for x in query:
                realisasi = realisasi + x.get_total_realisasi()
            
            total_ai_a = total_ai_a + sum_ai
            total_aki_a = total_aki_a + sum_aki
            total_realisasi_a = total_realisasi_a + realisasi

            sisa_aki = sum_aki - realisasi
            pct = (safe_div(realisasi,sum_aki))*100
            
            A_list.append((A_PRK_1[idx],data,sum_ai,sum_aki,realisasi,sisa_aki,pct,count))
        
        context["A_list"] = A_list
        context["count_a_total"] = PRKData.objects.filter(prk__kode_prk__in=A_PRK_2, file_lrpa=last_lrpa).count()
        context["total_ai_a"] = total_ai_a
        context["total_aki_a"] = total_aki_a
        context["total_realisasi_a"] = total_realisasi_a
        context["total_sisa_aki_a"] = total_aki_a - total_realisasi_a
        context["total_pct_a"] = (safe_div(total_realisasi_a,total_aki_a))*100
        
        #END COUNT FOR ALL PRK IN "Pekerjaan. Prasarana"

        #COUNT FOR ALL PRK IN "Pekerjaan. Utama"
        B_PRK_1 = ["Pembangunan KIT/TL/GI ", "Pengadaan KIT/TL/GI/Trafo", "Jasa konsultasi Enginering dan Konstruksi (KIT)", "Pemantauan & Pengelolaan Lingkungan Thp Kons", "Pengadaan Tower/Konduktor/MTU/HV/Trafo (TL dan GI)","Pek. Penyempurnaan KIT/Biaya Pra COD","Pek. Penyempurnaan TL dan GI"]
        B_PRK_2 = ["PB", "PD", "JS", "PL", "SWA","PP KIT","PP TL/GI"]

        B_list = []

        total_ai_b = 0
        total_aki_b = 0
        total_realisasi_b = 0

        count_b_total = 0
        
        for idx,data in enumerate(B_PRK_2):
            query = PRKData.objects.filter(prk__kode_prk=data, file_lrpa=last_lrpa)
            count = query.count()
            sum_ai = query.aggregate(Sum('ai_this_year'))['ai_this_year__sum']
            sum_aki = query.aggregate(Sum('aki_this_year'))['aki_this_year__sum']
            if not sum_ai: sum_ai = 0
            if not sum_aki: sum_aki = 0
            realisasi = 0
            for x in query:
                realisasi = realisasi + x.get_total_realisasi()
            
            total_ai_b = total_ai_b + sum_ai
            total_aki_b = total_aki_b + sum_aki
            total_realisasi_b = total_realisasi_b + realisasi

            sisa_aki = sum_aki - realisasi
            pct = (safe_div(realisasi,sum_aki))*100
            
            B_list.append((B_PRK_1[idx],data,sum_ai,sum_aki,realisasi,sisa_aki,pct,count))
        
        context["B_list"] = B_list
        context["count_b_total"] = PRKData.objects.filter(prk__kode_prk__in=B_PRK_2, file_lrpa=last_lrpa).count()
        context["total_ai_b"] = total_ai_b
        context["total_aki_b"] = total_aki_b
        context["total_realisasi_b"] = total_realisasi_b
        context["total_sisa_aki_b"] = total_aki_b - total_realisasi_b
        context["total_pct_b"] = (safe_div(total_realisasi_b,total_aki_b))*100

        #END COUNT FOR ALL PRK IN "Pekerjaan. Utama"

        #COUNT FOR ALL PRK IN "Pekerjaan. Lainnya"
        C_PRK_1 = ["IPPKH", "Jasa Bantuan Hukum & Manajemen Stakholder", "Jasa terkait SLO/Kommisioning (KIT)", "Jasa terkait SLO/Kommisioning (TL dan GI)", "Jasa Konstruksi & QA/QC oleh PMK (KIT)","Jasa Konstruksi & QA/QC oleh PMK (TL dan GI)","Biaya Sertifikat","Jasa Penyusunan Bid Doc dan HPE"]
        C_PRK_2 = ["KH", "KUM", "SER KIT", "SER TL/GI", "PMK KIT","PMK TL/GI","BS","PLIS"]

        C_list = []

        total_ai_c = 0
        total_aki_c = 0
        total_realisasi_c = 0

        count_c_total = 0

        for idx,data in enumerate(C_PRK_2):
            query = PRKData.objects.filter(prk__kode_prk=data, file_lrpa=last_lrpa)
            count = query.count()
            sum_ai = query.aggregate(Sum('ai_this_year'))['ai_this_year__sum']
            sum_aki = query.aggregate(Sum('aki_this_year'))['aki_this_year__sum']
            if not sum_ai: sum_ai = 0
            if not sum_aki: sum_aki = 0
            realisasi = 0
            for x in query:
                realisasi = realisasi + x.get_total_realisasi()
            
            total_ai_c = total_ai_c + sum_ai
            total_aki_c = total_aki_c + sum_aki
            total_realisasi_c = total_realisasi_c + realisasi

            sisa_aki = sum_aki - realisasi
            pct = (safe_div(realisasi,sum_aki))*100
            
            C_list.append((C_PRK_1[idx],data,sum_ai,sum_aki,realisasi,sisa_aki,pct,count))

        context["C_list"] = C_list
        context["count_c_total"] = PRKData.objects.filter(prk__kode_prk__in=C_PRK_2, file_lrpa=last_lrpa).count()
        context["total_ai_c"] = total_ai_c
        context["total_aki_c"] = total_aki_c
        context["total_realisasi_c"] = total_realisasi_c
        context["total_sisa_aki_c"] = total_aki_c - total_realisasi_c
        context["total_pct_c"] = (safe_div(total_realisasi_c,total_aki_c))*100

        total_ai_prk = total_ai_a + total_ai_b + total_ai_c
        total_aki_prk = total_aki_a + total_aki_b + total_aki_c
        total_realisasi_prk = total_realisasi_a + total_realisasi_b + total_realisasi_c
        total_sisa_aki_prk = total_aki_prk - total_realisasi_prk
        total_count_prk = context["count_a_total"] + context["count_b_total"] + context["count_c_total"]

        context["total_count_prk"] = total_count_prk
        context["total_ai_prk"] = total_ai_prk
        context["total_aki_prk"] = total_aki_prk
        context["total_realisasi_prk"] = total_realisasi_prk
        context["total_sisa_aki_prk"] = total_sisa_aki_prk
        context["total_pct_prk"] = (safe_div(total_realisasi_prk,total_aki_prk))*100
        #END COUNT FOR ALL PRK IN "Pekerjaan. Lainnya"

        return render(request, 'monev/monev_view.html', context)

class LKAIViewCOPY(LoginRequiredMixin, View):

    def get(self, request):
        context = {}
        context["month"] = this_month()
        # ALL DOCUMENT NEEDED
        if is_production():
            skai_1 = DocSKAI.objects.get(pk=1)
            skai_2 = DocSKAI.objects.get(pk=3)
            skai_3 = DocSKAI.objects.get(pk=6)
        else:
            skai_1 = DocSKAI.objects.get(pk=8)
            skai_2 = DocSKAI.objects.get(pk=10)
            skai_3 = DocSKAI.objects.get(pk=19)
        
        last_lrpa = get_last_lrpa()
        last_mou = FileMouPengalihan.objects.order_by('file_export_date').first()

        sq_1 = MacroData.objects.filter(macro_file=skai_1.macro.macro_file_1, prk=OuterRef('prk'))
        sq_2 = MacroData.objects.filter(macro_file=skai_2.macro.macro_file_1, prk=OuterRef('prk'))
        sq_3 = MacroData.objects.filter(macro_file=skai_3.macro.macro_file_1, prk=OuterRef('prk'))

        #GET USER DIVISION, DETERMINE USER VIEW
        division = request.user.division

        if division == "Super Admin" or division == "ANG":
            monitoring = get_all_prk_last_lrpa()
            context["for_div"] = "ALL"
        else:
            monitoring = get_all_prk_last_lrpa(rekap_user_induk=division)
            context["for_div"] = division
        
        lrpa = monitoring. \
               annotate(ai_1 = Round(Subquery(sq_1.values('ai_this_year')[:1])*1000), aki_1 = Round(Subquery(sq_1.values('aki_this_year')[:1])*1000), status_1 = sq_1.values('ang_status'),
               ai_2 = Round(Subquery(sq_2.values('ai_this_year')[:1])*1000), aki_2 = Round(Subquery(sq_2.values('aki_this_year')[:1])*1000), status_2 = sq_2.values('ang_status'),
               ai_3 = Round(Subquery(sq_3.values('ai_this_year')[:1])*1000), aki_3 = Round(Subquery(sq_3.values('aki_this_year')[:1])*1000), status_3 = sq_3.values('ang_status'),
               sd_1 = sq_1.values('sumber_dana'), sd_2 = sq_2.values('sumber_dana'), sd_3 = sq_3.values('sumber_dana'),
               )
        
        document = [skai_1, skai_2, last_lrpa, skai_3, last_mou]
        context["document"] = document #OPTIMIZE LATER? #YES USING RELATION IN PRK DATA!

        context["lrpa"] = lrpa

        for data in lrpa:
            print(data.prk.no_prk, data.get_total_realisasi(), data.is_aki_completed())

        return render(request, 'monev/monev_lkai_copy.html', context)

class LKAIView(LoginRequiredMixin, View): #DELETE LATER WITH TEMPLATE

    def get(self, request):
        context = {}
        context["month"] = this_month()
        #Total: 1.57s Python: 1.26s DB: 0.31s Queries: 14 

        # ALL DOCUMENT NEEDED
        if is_production():
            skai_1 = DocSKAI.objects.get(pk=1)
            skai_2 = DocSKAI.objects.get(pk=3)
            skai_3 = DocSKAI.objects.get(pk=6)
        else:
            skai_1 = DocSKAI.objects.get(pk=8)
            skai_2 = DocSKAI.objects.get(pk=10)
            skai_3 = DocSKAI.objects.get(pk=19)
        
        last_lrpa = LRPA_File.objects.order_by('-file_export_date').first()
        last_mou = FileMouPengalihan.objects.order_by('file_export_date').first()

        sq_1 = MacroData.objects.filter(macro_file=skai_1.macro.macro_file_1, prk=OuterRef('prk'))
        sq_2 = MacroData.objects.filter(macro_file=skai_2.macro.macro_file_1, prk=OuterRef('prk'))
        sq_3 = MacroData.objects.filter(macro_file=skai_3.macro.macro_file_1, prk=OuterRef('prk'))
        sq_mou = MouPengalihanData.objects.filter(file=last_mou, prk=OuterRef('prk'))

        #GET USER DIVISION, DETERMINE USER VIEW
        division = request.user.division

        if division == "Super Admin" or division == "ANG":
            monitoring = PRKData.objects.select_related('prk').filter(file=last_lrpa)
            context["for_div"] = "ALL"
        else:
            monitoring = PRKData.objects.select_related('prk').filter(file=last_lrpa, prk__rekap_user_induk=division)
            context["for_div"] = division
        
        lrpa = monitoring. \
               annotate(ai_1 = Round(Subquery(sq_1.values('ai_this_year')[:1])*1000), aki_1 = Round(Subquery(sq_1.values('aki_this_year')[:1])*1000), status_1 = sq_1.values('ang_status'),
               ai_2 = Round(Subquery(sq_2.values('ai_this_year')[:1])*1000), aki_2 = Round(Subquery(sq_2.values('aki_this_year')[:1])*1000), status_2 = sq_2.values('ang_status'),
               ai_3 = Round(Subquery(sq_3.values('ai_this_year')[:1])*1000), aki_3 = Round(Subquery(sq_3.values('aki_this_year')[:1])*1000), status_3 = sq_3.values('ang_status'),
               mou_jan = sq_mou.values('jan'),mou_feb = sq_mou.values('feb'),mou_mar = sq_mou.values('mar'),mou_apr = sq_mou.values('apr'),
               mou_mei = sq_mou.values('mei'),mou_jun = sq_mou.values('jun'),mou_jul = sq_mou.values('jul'),mou_aug = sq_mou.values('aug'),
               mou_sep = sq_mou.values('sep'),mou_okt = sq_mou.values('okt'),mou_nov = sq_mou.values('nov'),mou_des = sq_mou.values('des'),
               sd_1 = sq_1.values('sumber_dana'), sd_2 = sq_2.values('sumber_dana'), sd_3 = sq_3.values('sumber_dana'),
               )
        
        document = [skai_1, skai_2, last_lrpa, skai_3, last_mou]
        context["document"] = document #OPTIMIZE LATER?

        context["lrpa"] = lrpa

        for data in lrpa:
            print(data.no_prk, data.prk.pk)

        return render(request, 'monev/monev_lkai.html', context)

class UploadLRPA(LoginRequiredMixin, View):
    
    def get(self, request, *args, **kwargs):
        context = {}

        return render(request, 'monev/upload_lrpa.html', context)
    
    def post(self, request, *args, **kwargs):
        context = {}
        print(request.FILES)
        file = request.FILES["lrpa_file"]
        
        wb = load_workbook(file)
        ws = wb['Monitoring LRPA']

        # COMPROMISE ADDED COL
        start_col = 2
        end_col = 43
        
        list_rows = [idx for idx,cell in enumerate(ws["C"]) if cell.value and idx >= 6]
        #print(list_rows)
        
        doc_form = LRPAFileForm(request.POST, request.FILES)
        if doc_form.is_valid():
            doc = doc_form.save(commit=False)
            doc.upload_by = request.user
            doc.save()
            # NEW WAYS INSERT DIRECTLY TO PRKData
            for rows in list_rows:
                row = [cell.value for cell in ws[rows][start_col:end_col+1]]

                no_prk = row[0]
                try:
                    prk = PRK.objects.get(no_prk=row[0])
                except PRK.DoesNotExist:
                    prk = PRK(no_prk=no_prk).save()
                    print("CREATED NEW PRK",no_prk)

                try:
                    prk_data = PRKData.objects.get(prk=prk, prk_data_year=2022) #QUICK FIX, USE THIS YEAR LATER
                    print("USING PRK DATA FOR",no_prk)
                except PRKData.DoesNotExist:
                    prk_data = PRKData(prk=prk, prk_data_year=2022)
                    prk_data.save()
                    prk_data = PRKData.objects.get(prk=prk, prk_data_year=2022) #QUICK FIX, USE THIS YEAR LATER
                    print("CREATED NEW PRK DATA", no_prk)
                
                try:
                    prk_data.file_lrpa = doc
                    prk_data.disburse_year_before = row[9]
                    prk_data.jan_rencana = row[18]
                    prk_data.jan_realisasi = row[19]
                    prk_data.feb_rencana = row[20]
                    prk_data.feb_realisasi = row[21]
                    prk_data.mar_rencana = row[22]
                    prk_data.mar_realisasi = row[23]
                    prk_data.apr_rencana = row[24]
                    prk_data.apr_realisasi = row[25]
                    prk_data.mei_rencana = row[26]
                    prk_data.mei_realisasi = row[27]
                    prk_data.jun_rencana = row[28]
                    prk_data.jun_realisasi = row[29]
                    prk_data.jul_rencana = row[30]
                    prk_data.jul_realisasi = row[31]
                    prk_data.aug_rencana = row[32]
                    prk_data.aug_realisasi = row[33]
                    prk_data.sep_rencana = row[34]
                    prk_data.sep_realisasi = row[35]
                    prk_data.okt_rencana = row[36]
                    prk_data.okt_realisasi = row[37]
                    prk_data.nov_rencana = row[38]
                    prk_data.nov_realisasi = row[39]
                    prk_data.des_rencana = row[40]
                    prk_data.des_realisasi = row[41]
                    prk_data.mekanisme_pembayaran = row[14]
                    prk_data.ai_this_year = row[10]
                    prk_data.aki_this_year = row[11]
                    prk_data.save()
                    print("Saved", no_prk)
                except Exception as e:
                    exception_type, exception_object, exception_traceback = sys.exc_info()
                    line_number = exception_traceback.tb_lineno
                    print(e, line_number, exception_type, exception_object)
        else:
            print(doc_form.errors)

        return render(request, 'monev/upload_lrpa.html', context)

class UploadMouPengalihan(LoginRequiredMixin, View):

    def get(self, request):
        context = {}

        return render(request, 'monev/upload_mou.html', context)
    
    def post(self, request):
        context = {}

        file = request.FILES["mou_file"]

        wb = load_workbook(file, data_only=True)
        ws = wb['All Pengalihan']

        start_col = 2
        end_col = 19

        list_rows = [idx for idx,cell in enumerate(ws["C"]) if cell.value and idx >= 8]
        print(list_rows[-1])
        x = list_rows[-1]
        x = x + 1
        list_rows.append(x)
        print(list_rows[-1]) #USE THIS METHOD IN MACROS TOO !!

        doc_form = MouFileForm(request.POST, request.FILES)
        if doc_form.is_valid():
            doc = doc_form.save(commit=False)
            doc.upload_by = request.user
            doc.save()
            
            for rows in list_rows:
                row = [cell.value for cell in ws[rows][start_col:end_col+1]]

                no_prk = row[0]
                try:
                    prk = PRK.objects.get(no_prk=row[0])
                except PRK.DoesNotExist:
                    prk = PRK(no_prk=no_prk).save()
                    print("CREATED NEW PRK",no_prk)

                try:
                    prk_data = PRKData.objects.get(prk=prk, prk_data_year=2022) #QUICK FIX, USE THIS YEAR LATER
                    print("USING PRK DATA FOR",no_prk)
                except PRKData.DoesNotExist:
                    prk_data = PRKData(prk=prk, prk_data_year=2022)
                    prk_data.save()
                    prk_data = PRKData.objects.get(prk=prk, prk_data_year=2022) #QUICK FIX, USE THIS YEAR LATER
                    print("CREATED NEW PRK DATA", no_prk)
                
                try:
                    prk_data.file_mou = doc
                    prk_data.jan_pengalihan = row[6]
                    prk_data.feb_pengalihan = row[7]
                    prk_data.mar_pengalihan = row[8]
                    prk_data.apr_pengalihan = row[9]
                    prk_data.mei_pengalihan = row[10]
                    prk_data.jun_pengalihan = row[11]
                    prk_data.jul_pengalihan = row[12]
                    prk_data.aug_pengalihan = row[13]
                    prk_data.sep_pengalihan = row[14]
                    prk_data.okt_pengalihan = row[15]
                    prk_data.nov_pengalihan = row[16]
                    prk_data.des_pengalihan = row[17]
                    
                    prk_data.save()
                    print("Saved", no_prk)
                except Exception as e:
                    exception_type, exception_object, exception_traceback = sys.exc_info()
                    line_number = exception_traceback.tb_lineno
                    print(e, line_number, exception_type, exception_object)
        else:
            print(doc_form.errors)


        return render(request, 'monev/upload_mou.html', context)


class FileMonevList(LoginRequiredMixin, View):

    def get(self, request):
        context = {}

        doc = FileMouPengalihan.objects.all().order_by('file_export_date')
        lrpa = LRPA_File.objects.all().order_by('file_export_date')
        context["docs"] = doc
        context["lrpa"] = lrpa

        return render(request, 'monev/file_monev.html', context)

class EditPRK(LoginRequiredMixin, View):
    def test_func(self):
        return self.request.user.is_admin or self.request.user.is_staff
    
    def get(self, request, pk, type, *args, **kwargs):
        context = {}

        prk = PRK.objects.get(pk=pk)
        context["data"] = prk
        
        if type == 1:
            return render(request, 'monev/snippets/modal_prk_1.html', context)
        
        if type == 2:
            return render(request, 'monev/snippets/modal_prk_2.html', context)
        
        if type == 3:
            return render(request, 'monev/snippets/modal_prk_3.html', context)

    def post(self, request, pk, type, *args, **kwargs):
        context = {}

        data = request.POST
        prk = PRK.objects.get(no_prk = data["no_prk"])

        if type == 1:
            prk.kode_prk = data["kode_prk"]
            prk.save()
            context["data"] = prk
            return render(request, 'monev/snippets/inline_td_1.html', context)
        
        if type == 2:
            prk.kode_bpo = data["kode_bpo"]
            prk.upp = data["upp"]
            prk.save()
            context["data"] = prk
            return render(request, 'monev/snippets/inline_td_2.html', context)
        
        if type == 3:
            prk.rekap_user_induk = data["rekap_user_induk"]
            prk.save()
            context["data"] = prk
            return render(request, 'monev/snippets/inline_td_3.html', context)

